from typing import List, Optional
from datetime import datetime
import csv
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.api import deps
from app.db.models import User, Contact, Organization
from app.db.database import get_db
from app.schemas.schemas import ContactCreate, ContactResponse, ContactListResponse, ContactBase

router = APIRouter()

@router.get("", response_model=ContactListResponse)
async def get_contacts(
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    List contacts for the current user's organization.
    """
    query = select(Contact).where(Contact.organization_id == current_user.organization_id).offset(skip).limit(limit)
    result = await db.execute(query)
    contacts = result.scalars().all()
    
    # Simple total count (can be optimized)
    total_query = select(Contact).where(Contact.organization_id == current_user.organization_id)
    total_result = await db.execute(total_query)
    total = len(total_result.scalars().all())
    
    return {"items": contacts, "total": total}

@router.post("", response_model=ContactResponse)
async def create_contact(
    contact_in: ContactCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Create a new contact.
    """
    # Normalize and validate phone number to E.164
    from app.core.phone_utils import normalize_phone_number, validate_ghana_number
    phone = normalize_phone_number(contact_in.phone_number)
    if not validate_ghana_number(phone):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Ghana phone number: {contact_in.phone_number}"
        )

    # Check if contact already exists for this org
    query = select(Contact).where(
        Contact.organization_id == current_user.organization_id,
        Contact.phone_number == phone
    )
    result = await db.execute(query)
    existing_contact = result.scalar_one_or_none()
    if existing_contact:
        # Upsert: Update names if provided
        updated = False
        if contact_in.first_name and existing_contact.first_name != contact_in.first_name:
            existing_contact.first_name = contact_in.first_name
            updated = True
        if contact_in.last_name and existing_contact.last_name != contact_in.last_name:
            existing_contact.last_name = contact_in.last_name
            updated = True
        
        if updated:
            db.add(existing_contact)
            await db.commit()
            await db.refresh(existing_contact)
        return existing_contact

    from app.core.phone_utils import detect_network
    network_provider, _ = detect_network(phone)

    db_obj = Contact(
        first_name=contact_in.first_name,
        last_name=contact_in.last_name,
        phone_number=phone,
        organization_id=current_user.organization_id,
        network=network_provider.value if network_provider else None
    )
    db.add(db_obj)
    await db.commit()
    await db.refresh(db_obj)
    return db_obj

@router.post("/upload")
async def upload_contacts(
    file: UploadFile = File(...),
    group_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Upload contacts from a CSV file.
    CSV header: first_name, last_name, phone_number
    If group_id is provided, both new and existing contacts in the CSV will be added to the group.
    """
    # Verify group if provided
    if group_id:
        from app.db.models import ContactGroup
        group_query = select(ContactGroup).where(
            ContactGroup.id == group_id,
            ContactGroup.organization_id == current_user.organization_id
        )
        group = (await db.execute(group_query)).scalar_one_or_none()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")

    content = await file.read()
    is_xlsx = file.filename.lower().endswith(('.xlsx', '.xls'))
    rows = []

    if is_xlsx:
        import io
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            sheet = wb.active
            
            # Read header row
            header_row = next(sheet.iter_rows(values_only=True), None)
            if not header_row:
                raise HTTPException(status_code=400, detail="Excel file is empty.")
            
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            
            # Read content rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # If row is completely empty, skip
                if not any(val is not None for val in row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = val
                rows.append(row_dict)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    else:
        try:
            decoded = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                decoded = content.decode('latin-1')
            except:
                raise HTTPException(status_code=400, detail="Could not decode file. Please use UTF-8 or Latin-1 encoding.")
                
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    
    created_count = 0
    skipped_count = 0
    group_added_count = 0
    
    from app.db.models import contact_group_association
    
    # We will accumulate contacts to add to the group
    contacts_to_add_to_group = []

    # Define header mappings (all lowercase — headers are lowercased before comparison)
    # Covers every realistic user-created column name variation
    header_map = {
        'first_name': [
            # "name" as a single column (treated as first name)
            'name',
            # first_name variations
            'first_name', 'first name', 'firstname', 'first-name',
            # fname variations
            'fname', 'f name', 'f_name',
            # given name
            'given name', 'given_name', 'givenname',
            # other common labels
            'forename', 'fore name', 'fore_name',
            'customer name', 'customer_name', 'customername',
            'client name', 'client_name', 'clientname',
            'full name', 'full_name', 'fullname',
        ],
        'last_name': [
            # last_name variations
            'last_name', 'last name', 'lastname', 'last-name',
            # lname variations
            'lname', 'l name', 'l_name',
            # surname variations
            'surname', 'sur name', 'sur_name',
            'family name', 'family_name', 'familyname',
            'second name', 'second_name', 'secondname',
        ],
        'phone_number': [
            # phone variations
            'phone', 'phone_number', 'phone number', 'phonenumber', 'phone-number',
            'phone no', 'phone_no', 'phoneno',
            'phone #', 'phone#',
            # contact number variations
            'contact', 'contact number', 'contact_number', 'contactnumber', 'contact-number',
            'contact no', 'contact_no', 'contactno',
            # mobile variations
            'mobile', 'mobile number', 'mobile_number', 'mobilenumber', 'mobile-number',
            'mobile no', 'mobile_no', 'mobileno',
            'mobile phone', 'mobile_phone', 'mobilephone',
            'cell', 'cell number', 'cell_number', 'cellnumber',
            'cell no', 'cell_no', 'cellno',
            'cellphone', 'cell phone', 'cell_phone',
            # telephone variations
            'tel', 'telephone', 'tel number', 'tel_number', 'telnumber',
            'telephone number', 'telephone_number', 'telephonenumber',
            # number variations
            'number', 'num', 'no',
            # other
            'msisdn', 'gsm', 'gsm number', 'gsm_number',
            'whatsapp', 'whatsapp number', 'whatsapp_number',
            'recipient', 'recipient number', 'recipient_number',
        ]
    }

    def get_value(row, internal_key):
        for header_key, value in row.items():
            if value is None:
                continue
            
            # Excel sometimes parses numbers as floats (e.g. 541234567.0 instead of 0541234567)
            if isinstance(value, float) and value.is_integer():
                value = int(value)
                
            header_lower = str(header_key).strip().lower()
            # Match directly or with spaces↔underscores normalized
            normalized = header_lower.replace(' ', '_').replace('-', '_')
            if header_lower in header_map[internal_key] or normalized in header_map[internal_key]:
                return value
                
        # Fallback to literal key
        val = row.get(internal_key)
        if isinstance(val, float) and val.is_integer():
            val = int(val)
        return val

    for row in rows:
        phone_raw = get_value(row, 'phone_number')
        if not phone_raw:
            skipped_count += 1
            continue

        # Normalize and validate phone number to E.164
        from app.core.phone_utils import normalize_phone_number, validate_ghana_number
        try:
            phone = normalize_phone_number(str(phone_raw).strip())
        except Exception:
            skipped_count += 1
            continue

        if not validate_ghana_number(phone):
            skipped_count += 1
            continue
        
        # Check if exists
        query = select(Contact).where(
            Contact.organization_id == current_user.organization_id,
            Contact.phone_number == phone
        )
        existing_contact = (await db.execute(query)).scalar_one_or_none()
        
        if existing_contact:
            skipped_count += 1
            if group_id:
                contacts_to_add_to_group.append(existing_contact.id)
            continue
            
        first_name_val = get_value(row, 'first_name')
        last_name_val = get_value(row, 'last_name')
        
        from app.core.phone_utils import detect_network
        network_provider, _ = detect_network(phone)

        db_obj = Contact(
            phone_number=phone,
            first_name=str(first_name_val).strip() if first_name_val is not None else None,
            last_name=str(last_name_val).strip() if last_name_val is not None else None,
            organization_id=current_user.organization_id,
            network=network_provider.value if network_provider else None
        )
        db.add(db_obj)
        await db.flush() # flush to get the id
        created_count += 1
        
        if group_id:
            contacts_to_add_to_group.append(db_obj.id)
            
    if group_id and contacts_to_add_to_group:
        # Check existing group members to avoid duplicates in association
        assoc_query = select(contact_group_association.c.contact_id).where(
            contact_group_association.c.group_id == group_id
        )
        existing_group_ids = {row[0] for row in (await db.execute(assoc_query)).all()}
        
        for cid in set(contacts_to_add_to_group):
            if cid not in existing_group_ids:
                await db.execute(
                    contact_group_association.insert().values(contact_id=cid, group_id=group_id)
                )
                group_added_count += 1

    await db.commit()
    return {
        "message": f"Successfully imported {created_count} contacts.",
        "created": created_count,
        "skipped": skipped_count,
        "group_added": group_added_count
    }

@router.delete("/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_contact(
    contact_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Delete a contact.
    """
    query = select(Contact).where(Contact.id == contact_id, Contact.organization_id == current_user.organization_id)
    result = await db.execute(query)
    db_obj = result.scalar_one_or_none()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    await db.delete(db_obj)
    await db.commit()
    return None

@router.patch("/{contact_id}", response_model=ContactResponse)
async def update_contact(
    contact_id: int,
    contact_in: ContactCreate, # Reusing ContactCreate for simplicity or create a ContactUpdate schema
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Update a contact.
    """
    query = select(Contact).where(Contact.id == contact_id, Contact.organization_id == current_user.organization_id)
    result = await db.execute(query)
    db_obj = result.scalar_one_or_none()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    update_data = contact_in.model_dump(exclude_unset=True)
    if "phone_number" in update_data:
        from app.core.phone_utils import normalize_phone_number, validate_ghana_number, detect_network
        normalized = normalize_phone_number(update_data["phone_number"])
        if not validate_ghana_number(normalized):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid Ghana phone number: {update_data['phone_number']}"
            )
        update_data["phone_number"] = normalized
        # Update the network provider
        network_provider, _ = detect_network(normalized)
        db_obj.network = network_provider.value if network_provider else None

    for key, value in update_data.items():
        setattr(db_obj, key, value)

    db.add(db_obj)
    await db.commit()
    await db.refresh(db_obj)
    return db_obj


@router.patch("/{contact_id}/opt-out", status_code=200)
async def opt_out_contact(
    contact_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Mark a contact as opted-out (STOP). No further messages will be sent to them.
    Can be triggered manually by an admin or automatically when a recipient replies STOP.
    """
    query = select(Contact).where(
        Contact.id == contact_id,
        Contact.organization_id == current_user.organization_id
    )
    result = await db.execute(query)
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    contact.is_opted_out = True
    contact.opted_out_at = datetime.utcnow()
    await db.commit()
    return {"message": f"Contact {contact.phone_number} has been opted out successfully."}


@router.patch("/{contact_id}/opt-in", status_code=200)
async def opt_in_contact(
    contact_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """
    Re-subscribe a contact who previously opted out.
    Only use when you have explicit renewed consent from the recipient.
    """
    query = select(Contact).where(
        Contact.id == contact_id,
        Contact.organization_id == current_user.organization_id
    )
    result = await db.execute(query)
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    contact.is_opted_out = False
    contact.opted_out_at = None
    await db.commit()
    return {"message": f"Contact {contact.phone_number} has been opted back in."}
